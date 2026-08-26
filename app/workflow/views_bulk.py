import csv
import io
import json
import re
import zipfile
from datetime import datetime

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from catalog.models import PackagingFile, Product

PAIR_SUFFIX_RE = re.compile(r'(_ai|_jpg|_jpeg|_png|_gif|_pdf)$', re.IGNORECASE)
VISUAL_EXTENSIONS = ('jpg', 'jpeg', 'png', 'gif', 'pdf')


@login_required
def bulk_home(request):
    """catalog.views.product_list와 같은 이유로 품목별 current_final_file()을 템플릿에서
    개별 호출하면 품목 수(현재 258건)만큼 쿼리가 늘어나 워커 타임아웃(502/500)으로
    이어진다 — 최종 승인본을 한 번에 조회해 품목별로 파이썬에서 묶어 전달한다.

    258건을 한 화면에 전부 그리면 스크롤이 너무 길어지므로, 검색(품목명·품목코드)과
    페이지네이션으로 좁혀볼 수 있게 한다 — 매 페이지마다 그 페이지에 보이는 품목의
    최종 승인본만 조회하므로 전체 건수가 늘어나도 쿼리 수는 늘지 않는다."""
    q = request.GET.get('q', '').strip()
    products = Product.objects.filter(is_active=True).order_by('name')
    if q:
        products = products.filter(Q(name__icontains=q) | Q(code__icontains=q))

    paginator = Paginator(products, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    product_ids = [p.pk for p in page_obj]
    final_file_by_product = {}
    for f in (PackagingFile.objects
              .filter(product_id__in=product_ids, is_active=True,
                      status=PackagingFile.Status.FINAL_APPROVED)
              .order_by('product_id', '-version')):
        final_file_by_product.setdefault(f.product_id, f)  # 품목별로 최신 버전 하나만 남김

    rows = [{'product': p, 'final_file': final_file_by_product.get(p.pk)} for p in page_obj]

    # 일괄 업로드 화면에서 파일 묶음마다 품목을 직접 골라 연결할 수 있도록, 검색/
    # 페이지네이션과 무관하게 활성 품목 전체를 가벼운 목록(id/code/name)으로 넘긴다 —
    # 258건 수준이면 페이지 하나에 실어도 부담 없다(현재 파일 목록과는 별개).
    all_products = list(
        Product.objects.filter(is_active=True).order_by('name').values('id', 'code', 'name'))

    return render(request, 'workflow/bulk.html', {
        'rows': rows, 'page_obj': page_obj, 'q': q, 'all_products': all_products,
    })


@login_required
def bulk_upload_history(request):
    """일괄 업로드는 특정 재발주 건(ReorderRequest)에 매이지 않는 품목 파일 관리라,
    그 건의 이력(타임라인)에는 남지 않는다 — 나중에 "이 버전이 왜 갑자기 바뀌었는지"
    추적할 곳이 없어지는 문제가 있어, 일괄 업로드로 등록된 버전만 모아 별도로 보여준다."""
    q = request.GET.get('q', '').strip()
    logs = (PackagingFile.objects
            .filter(is_bulk_upload=True)
            .select_related('product', 'uploaded_by', 'approved_by')
            .order_by('-uploaded_at'))
    if q:
        logs = logs.filter(Q(product__name__icontains=q) | Q(product__code__icontains=q) | Q(note__icontains=q))

    paginator = Paginator(logs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'workflow/bulk_upload_history.html', {'page_obj': page_obj, 'q': q})


@login_required
def bulk_mapping_template(request):
    """일괄 업로드용 빈 엑셀 매핑표 양식을 다운로드한다(_read_mapping이 읽는 형식과 동일)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '매핑표'

    header = ['파일명', '품목코드', '품목명', '승인일', '승인자', '비고']
    ws.append(header)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    ws.append(['예시_그린비료_v1.ai', 'FERT-PP-1001', '그린비료 20kg PP포대', '2026-01-15', '조현종', '예시 행입니다 — 지우고 실제 데이터를 입력하세요'])
    for cell in ws[2]:
        cell.font = openpyxl.styles.Font(italic=True, color='888888')

    widths = [26, 14, 26, 12, 10, 34]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mapping_template.xlsx"'
    return response


def _read_mapping(mapping_file):
    """엑셀 매핑표(파일명/품목코드/품목명/승인일/승인자/비고) → {파일명(확장자 제외): row dict}"""
    mapping = {}
    wb = openpyxl.load_workbook(mapping_file, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return mapping
    header = [str(h).strip() if h else '' for h in rows[0]]
    for row in rows[1:]:
        data = dict(zip(header, row))
        filename = str(data.get('파일명') or '').strip()
        if not filename:
            continue
        key = PAIR_SUFFIX_RE.sub('', filename.rsplit('.', 1)[0])
        mapping[key] = {
            'item_code': str(data.get('품목코드') or '').strip(),
            'product_name': str(data.get('품목명') or '').strip(),
            'approved_date': data.get('승인일'),
            'approver': str(data.get('승인자') or '').strip(),
            'note': str(data.get('비고') or '').strip(),
        }
    return mapping


def _resolve_product(base_key, map_row, product_id=None):
    """화면에서 직접 품목을 골랐으면(product_id) 그걸로 확정 — 엑셀 매핑표나 파일명
    추정은 아예 건너뛴다. product_id가 없을 때만 기존 방식(품목코드 upsert 또는
    품목명 매칭)으로 넘어간다."""
    if product_id:
        return Product.objects.filter(pk=product_id).first()

    item_code = map_row.get('item_code') if map_row else ''
    product_name = (map_row.get('product_name') if map_row else '') or base_key

    if item_code:
        product = Product.objects.filter(code=item_code).first()
        if product:
            if product_name and product.name != product_name:
                product.name = product_name
                product.save(update_fields=['name'])
            return product
        if map_row.get('product_name'):
            # 매핑표에 유형/제품군 정보가 없으므로 기본값으로 생성 — 필요 시 /admin에서 보정.
            return Product.objects.create(
                code=item_code, name=product_name,
                category=Product.Category.LABEL, product_line=Product.ProductLine.FERTILIZER,
            )
        return None

    product = Product.objects.filter(name=product_name).first()
    if not product:
        product = next((p for p in Product.objects.all() if p.name in base_key), None)
    return product


@login_required
def bulk_upload(request):
    if request.method != 'POST':
        return redirect('workflow:bulk')

    files = request.FILES.getlist('files')
    mapping_file = request.FILES.get('mapping_file')
    excel_mapping = _read_mapping(mapping_file) if mapping_file else {}

    # 화면에서 파일 묶음별로 직접 고른 품목/승인 정보 — bulk.html의 JS가 파일을
    # 고르는 즉시 같은 규칙(PAIR_SUFFIX_RE)으로 그룹을 만들어 미리 보여주고, 제출
    # 시점에 이 JSON 하나로 실어 보낸다. 매번 엑셀 매핑표를 새로 만들 필요 없이
    # 화면에서 바로 연결할 수 있게 하기 위함 — 엑셀 매핑표는 대량 이관 등에서
    # 여전히 쓸 수 있도록 남겨두되, 화면 선택이 있으면 그쪽을 우선한다.
    try:
        ui_mapping = json.loads(request.POST.get('mapping_json') or '{}')
    except ValueError:
        ui_mapping = {}

    groups = {}
    for f in files:
        stem, _, ext = f.name.rpartition('.')
        base_key = PAIR_SUFFIX_RE.sub('', stem or f.name)
        groups.setdefault(base_key, {})[ext.lower()] = f

    registered, unmatched = [], []
    for base_key, pair in groups.items():
        ai_file = pair.get('ai')
        jpg_file = next((pair[ext] for ext in VISUAL_EXTENSIONS if ext in pair), None)
        if not ai_file or not jpg_file:
            unmatched.append(f'{base_key} (AI/이미지·PDF 짝이 맞지 않음)')
            continue

        ui_row = ui_mapping.get(base_key)
        if ui_row and ui_row.get('product_id'):
            map_row = {
                'note': ui_row.get('note', ''),
                'approver': ui_row.get('approver', ''),
                'approved_date': ui_row.get('approved_date') or None,
            }
            product = _resolve_product(base_key, map_row, product_id=ui_row['product_id'])
        else:
            map_row = excel_mapping.get(base_key)
            product = _resolve_product(base_key, map_row)
        if not product:
            unmatched.append(f'{base_key} (품목 미매칭 — 화면에서 품목을 선택하거나 엑셀 매핑표에 품목코드·품목명을 지정하세요)')
            continue

        pkg = PackagingFile.objects.create(
            product=product, ai_file=ai_file, jpg_file=jpg_file, uploaded_by=request.user,
            note=(map_row['note'] if map_row else '') or '일괄 이관 업로드',
            is_bulk_upload=True,
        )
        if map_row and map_row.get('approver'):
            approved_at = None
            raw_date = map_row.get('approved_date')
            if isinstance(raw_date, datetime):
                approved_at = raw_date
            elif raw_date:
                try:
                    approved_at = datetime.strptime(str(raw_date), '%Y-%m-%d')
                except ValueError:
                    approved_at = None
            pkg.approve(request.user)
            if approved_at:
                if timezone.is_naive(approved_at):
                    approved_at = timezone.make_aware(approved_at)
                pkg.approved_at = approved_at
            pkg.note = f"{pkg.note} · 원 승인자: {map_row['approver']}"
            pkg.save(update_fields=['approved_at', 'note'])
        registered.append(f'{product.name} v{pkg.version}')

    if registered:
        messages.success(request, f'{len(registered)}건 등록됨: ' + ', '.join(registered))
    if unmatched:
        messages.warning(request, f'{len(unmatched)}건 미매칭(등록되지 않음): ' + '; '.join(unmatched))
    if not registered and not unmatched:
        messages.error(request, '업로드할 파일을 선택해주세요.')
    return redirect('workflow:bulk')


@login_required
def bulk_download(request):
    if request.method != 'POST':
        return redirect('workflow:bulk')

    product_ids = request.POST.getlist('products')
    products = Product.objects.filter(pk__in=product_ids)
    if not products:
        messages.error(request, '다운로드할 품목을 하나 이상 선택해주세요.')
        return redirect('workflow:bulk')

    buffer = io.BytesIO()
    manifest_rows = [['품목코드', '품목명', '버전', '승인일', '승인자', '파일']]
    skipped = []
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for product in products:
            final_file = product.current_final_file()
            if not final_file:
                skipped.append(f'{product.code} {product.name}')
                continue
            folder = re.sub(r'[\\/:*?"<>|]', '_', f'{product.code}_{product.name}')
            for field, display_name in (
                (final_file.ai_file, final_file.ai_display_filename),
                (final_file.jpg_file, final_file.jpg_display_filename),
            ):
                if field:
                    # 개별 다운로드(품목명_버전_날짜.확장자)와 같은 파일명 규칙을
                    # 압축 파일 안에서도 그대로 적용한다.
                    arcname = f'{folder}/{display_name}'
                    field.open('rb')
                    zf.writestr(arcname, field.read())
                    field.close()
                    manifest_rows.append([
                        product.code, product.name, final_file.version,
                        final_file.approved_at.strftime('%Y-%m-%d') if final_file.approved_at else '',
                        final_file.approved_by.get_full_name() if final_file.approved_by else '',
                        arcname,
                    ])
        manifest_io = io.StringIO()
        csv.writer(manifest_io).writerows(manifest_rows)
        # UTF-8 BOM 없이 저장하면 Excel(특히 한글 Windows)이 CP949로 잘못 인식해
        # 한글이 깨져 보인다 — utf-8-sig로 BOM을 붙여줘야 Excel에서 바로 정상 표시된다.
        zf.writestr('manifest.csv', manifest_io.getvalue().encode('utf-8-sig'))

    if skipped:
        messages.warning(request, f'승인본이 없어 제외된 품목: {", ".join(skipped)}')

    buffer.seek(0)
    # 헤더는 ASCII만 허용되므로 파일명은 영문으로 구성 (내용물 파일명은 한글 유지).
    filename = f'packaging_final_files_{datetime.now().strftime("%Y%m%d_%H%M")}.zip'
    response = HttpResponse(buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
